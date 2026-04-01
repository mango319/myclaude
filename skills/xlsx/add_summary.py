from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import sys
sys.stdout.reconfigure(encoding='utf-8')

file_path = r'C:\Users\Administrator\Desktop\小篆便携式打印机OEM定制日程及费用.xlsx'
wb = load_workbook(file_path)
sheet = wb['OEM定制研发费用表']

# 定义样式
title_font = Font(bold=True, size=12, color='FFFFFF')
header_font = Font(bold=True, size=11)
border_style = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
title_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
subtitle_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

# 在第11行（报价说明前）插入汇总区域
insert_row = 11

# 插入行
sheet.insert_rows(insert_row, 10)

# 设置列宽
sheet.column_dimensions['A'].width = 5
sheet.column_dimensions['B'].width = 25
sheet.column_dimensions['C'].width = 20
sheet.column_dimensions['D'].width = 12
sheet.column_dimensions['E'].width = 8
sheet.column_dimensions['F'].width = 15
sheet.column_dimensions['G'].width = 25

# === 添加汇总标题 ===
current_row = insert_row
cell = sheet.cell(current_row, 2)
cell.value = '三、费用及周期汇总'
cell.font = Font(bold=True, size=13)
cell.alignment = Alignment(horizontal='center', vertical='center')
sheet.merge_cells(f'B{current_row}:G{current_row}')
current_row += 1

# === 研发费用汇总 ===
cell = sheet.cell(current_row, 2)
cell.value = '1、研发定制费用'
cell.font = header_font
cell.fill = subtitle_fill
cell.alignment = Alignment(horizontal='center', vertical='center')
sheet.merge_cells(f'B{current_row}:C{current_row}')
current_row += 1

# 表头
headers = ['费用项目', '金额(万元)', '完成周期']
for col_idx, header in enumerate(headers, start=2):
    cell = sheet.cell(current_row, col_idx)
    cell.value = header
    cell.font = title_font
    cell.fill = title_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border_style
current_row += 1

# 数据行 - 研发费用
# 设备端
sheet.cell(current_row, 2).value = '打印机设备端定制费'
sheet.cell(current_row, 3).value = '=F4'
sheet.cell(current_row, 4).value = '3周'
for col in range(2, 5):
    sheet.cell(current_row, col).border = border_style
    sheet.cell(current_row, col).alignment = Alignment(horizontal='center' if col > 2 else 'left', vertical='center')
current_row += 1

# PC驱动
sheet.cell(current_row, 2).value = 'PC驱动软件定制费'
sheet.cell(current_row, 3).value = '=F5'
sheet.cell(current_row, 4).value = '4周'
for col in range(2, 5):
    sheet.cell(current_row, col).border = border_style
    sheet.cell(current_row, col).alignment = Alignment(horizontal='center' if col > 2 else 'left', vertical='center')
current_row += 1

# App软件
sheet.cell(current_row, 2).value = 'App软件定制费'
sheet.cell(current_row, 3).value = '=F6'
sheet.cell(current_row, 4).value = '4周'
for col in range(2, 5):
    sheet.cell(current_row, col).border = border_style
    sheet.cell(current_row, col).alignment = Alignment(horizontal='center' if col > 2 else 'left', vertical='center')
current_row += 1

# 微信小程序
sheet.cell(current_row, 2).value = '微信小程序定制费'
sheet.cell(current_row, 3).value = '=F7'
sheet.cell(current_row, 4).value = '4周'
for col in range(2, 5):
    sheet.cell(current_row, col).border = border_style
    sheet.cell(current_row, col).alignment = Alignment(horizontal='center' if col > 2 else 'left', vertical='center')
current_row += 1

# 税费
sheet.cell(current_row, 2).value = '税费(6%)'
sheet.cell(current_row, 3).value = '=F9'
sheet.cell(current_row, 4).value = '-'
for col in range(2, 5):
    sheet.cell(current_row, col).border = border_style
    sheet.cell(current_row, col).alignment = Alignment(horizontal='center' if col > 2 else 'left', vertical='center')
current_row += 1

# 研发费用小计
sheet.cell(current_row, 2).value = '研发费用小计'
sheet.cell(current_row, 2).font = Font(bold=True)
sheet.cell(current_row, 3).value = '=SUM(F4:F7)+F9'
sheet.cell(current_row, 3).font = Font(bold=True)
sheet.cell(current_row, 4).value = '4周'
sheet.cell(current_row, 4).font = Font(bold=True)
for col in range(2, 5):
    sheet.cell(current_row, col).border = border_style
    sheet.cell(current_row, col).alignment = Alignment(horizontal='center' if col > 2 else 'left', vertical='center')
    sheet.cell(current_row, col).fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
current_row += 1

# === 认证费用汇总 ===
current_row += 1
cell = sheet.cell(current_row, 2)
cell.value = '2、认证费用'
cell.font = header_font
cell.fill = subtitle_fill
cell.alignment = Alignment(horizontal='center', vertical='center')
sheet.merge_cells(f'B{current_row}:C{current_row}')
current_row += 1

# 表头
headers = ['认证项目', '费用(万元)', '认证时间', '说明']
for col_idx, header in enumerate(headers, start=2):
    cell = sheet.cell(current_row, col_idx)
    cell.value = header
    cell.font = title_font
    cell.fill = title_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border_style
current_row += 1

# CCC认证
sheet.cell(current_row, 2).value = 'CCC强制认证'
sheet.cell(current_row, 3).value = 0.4
sheet.cell(current_row, 4).value = '2周'
sheet.cell(current_row, 5).value = '可派生'
for col in range(2, 6):
    sheet.cell(current_row, col).border = border_style
    sheet.cell(current_row, col).alignment = Alignment(horizontal='center' if col > 2 else 'left', vertical='center')
current_row += 1

# SRRC认证
sheet.cell(current_row, 2).value = 'SRRC无线入网认证'
sheet.cell(current_row, 3).value = 0.4
sheet.cell(current_row, 4).value = '3~4月'
sheet.cell(current_row, 5).value = 'Wi-Fi机型需要'
for col in range(2, 6):
    sheet.cell(current_row, col).border = border_style
    sheet.cell(current_row, col).alignment = Alignment(horizontal='center' if col > 2 else 'left', vertical='center')
current_row += 1

# 能效认证
sheet.cell(current_row, 2).value = '能效认证'
sheet.cell(current_row, 3).value = 0.3
sheet.cell(current_row, 4).value = '2周'
sheet.cell(current_row, 5).value = '必须'
for col in range(2, 6):
    sheet.cell(current_row, col).border = border_style
    sheet.cell(current_row, col).alignment = Alignment(horizontal='center' if col > 2 else 'left', vertical='center')
current_row += 1

# 认证费用小计
sheet.cell(current_row, 2).value = '认证费用小计'
sheet.cell(current_row, 2).font = Font(bold=True)
sheet.cell(current_row, 3).value = '=SUM(D' + str(current_row-3) + ':D' + str(current_row-1) + ')'
sheet.cell(current_row, 3).font = Font(bold=True)
sheet.cell(current_row, 4).value = '-'
sheet.cell(current_row, 4).font = Font(bold=True)
for col in range(2, 5):
    sheet.cell(current_row, col).border = border_style
    sheet.cell(current_row, col).alignment = Alignment(horizontal='center' if col > 2 else 'left', vertical='center')
    sheet.cell(current_row, col).fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
current_row += 1

# === 总计 ===
current_row += 1
# 先设置值，再合并
cell = sheet.cell(current_row, 2)
cell.value = '3、项目总计'
cell.font = Font(bold=True, size=12, color='FFFFFF')
cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
cell.alignment = Alignment(horizontal='center', vertical='center')
sheet.merge_cells(f'B{current_row}:C{current_row}')
current_row += 1

# 表头
headers = ['项目', '金额(万元)', '周期']
for col_idx, header in enumerate(headers, start=2):
    cell = sheet.cell(current_row, col_idx)
    cell.value = header
    cell.font = title_font
    cell.fill = title_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border_style
current_row += 1

# 研发费用总计
research_row = current_row
sheet.cell(current_row, 2).value = '研发定制费用总计'
sheet.cell(current_row, 2).font = Font(bold=True)
sheet.cell(current_row, 3).value = '=D' + str(current_row - 14)
sheet.cell(current_row, 3).font = Font(bold=True)
sheet.cell(current_row, 4).value = '4周'
for col in range(2, 5):
    sheet.cell(current_row, col).border = border_style
    sheet.cell(current_row, col).alignment = Alignment(horizontal='center' if col > 2 else 'left', vertical='center')
current_row += 1

# 认证费用总计
cert_row = current_row
sheet.cell(current_row, 2).value = '认证费用总计'
sheet.cell(current_row, 2).font = Font(bold=True)
sheet.cell(current_row, 3).value = '=D' + str(current_row - 6)
sheet.cell(current_row, 3).font = Font(bold=True)
sheet.cell(current_row, 4).value = '2周'
for col in range(2, 5):
    sheet.cell(current_row, col).border = border_style
    sheet.cell(current_row, col).alignment = Alignment(horizontal='center' if col > 2 else 'left', vertical='center')
current_row += 1

# 总计
sheet.cell(current_row, 2).value = '项目费用总计'
sheet.cell(current_row, 2).font = Font(bold=True, size=12)
sheet.cell(current_row, 3).value = '=D' + str(research_row) + '+D' + str(cert_row)
sheet.cell(current_row, 3).font = Font(bold=True, size=12)
sheet.cell(current_row, 3).number_format = '#,##0.00'
sheet.cell(current_row, 4).value = '4~6月'
sheet.cell(current_row, 4).font = Font(bold=True, size=12)
for col in range(2, 5):
    sheet.cell(current_row, col).border = border_style
    sheet.cell(current_row, col).alignment = Alignment(horizontal='center' if col > 2 else 'left', vertical='center')
    sheet.cell(current_row, col).fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')

# 保存文件
wb.save(file_path)
print('文件已更新，添加了费用及周期汇总')
